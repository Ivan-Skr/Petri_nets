from flask import Flask, render_template, json, request, redirect, url_for, send_from_directory
import os
import multiprocessing
import openpyxl


from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from database_setup import Base, Task, User


UPLOAD_FOLDER = 'upload_files'
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

engine = create_engine('sqlite:///main.db?check_same_thread=False')
Base.metadata.bind = engine

DBSession = sessionmaker(bind=engine)
session = DBSession()


@app.route('/')
@app.route('/index')
def index():
    global alert
    alert = 0
    return render_template('index.html')


@app.route('/training')
def instruction():
    global alert
    alert = 0
    return render_template('instruction.html')


@app.route('/simulator')
def simulator():
    global alert
    alert = 0
    return render_template('simulator.html')


@app.route('/theory')
def theory():
    global alert
    alert = 0
    return render_template('theory.html')


alert = 0

@app.route('/<name>/delete/', methods=['POST', 'GET'])
def delete(name):
    global alert
    whitelist = ['User1']                                          ######################################## Whitelist
    deleteUser = session.query(User).filter_by(name=name).one()
    if request.method == 'POST':
        if os.environ.get( "USERNAME" ) in whitelist:
            session.delete(deleteUser)
            session.commit()
            alert = 0
            return redirect(url_for('results'))
        else:
            alert = 1
            return redirect(url_for('results'))
    else:
        return render_template('delete.html', name=name)


@app.route('/results')
def results():
    global alert
    users = session.query(User).all()
    us = []
    for _ in users:
        us.append((_.id, _.name, _.result, _.result_procent, _.username))
    us.sort(key=lambda x: x[2], reverse=True)

    users_updated = []
    for _ in range(len(us)):
        users_updated.append(User(name=us[_][1], result=us[_][2], result_procent=us[_][3], username=us[_][4]))


    workbook = openpyxl.load_workbook(filename="upload_files\petri_nets_results.xlsx")
    sheet = workbook.worksheets[0]
    sheet.delete_cols(1, 6)
    
    
    for _ in range(1, len(us)+1):
        for _1 in range(1, 5):
            if _1 == 3:
                sheet.cell(row=_, column=_1, value=str(us[_-1][_1])+'%')
            else:
                sheet.cell(row=_, column=_1, value=us[_-1][_1])


    workbook.save("upload_files\petri_nets_results.xlsx")

    return render_template('results.html', users = users_updated, lenght = tasks_lenght, alert=alert)


@app.route('/' + UPLOAD_FOLDER + '/<file_name>')
def download(file_name):
    return send_from_directory(app.config['UPLOAD_FOLDER'], file_name, as_attachment=True)


name = ''
result = 0
get_tasks = session.query(Task).all()
tasks_lenght = len(get_tasks)


@app.route('/tasks/<task_num>/', methods=['POST', 'GET'])
def tasks(task_num):
    global name, result, alert
    alert = 0
    if request.method == 'POST':
        if task_num == '0':
            name = ''
            result = 0
            name = request.form['name']

            return redirect(url_for("tasks", task_num=1))
        else:
            get_tasks = session.query(Task).all()
            task = get_tasks[int(task_num)-1]
            listt = list(range(0, task.iterations+1))
            
            answers = []
            for _ in range(int(task.iterations)):
                a = []
                for _1 in range(int(task.sets)):
                    a.append(int(request.form[f'class{_+2}-{_1+1}']))
                answers.append(a)

            
            if str(answers) == (task.answers):
                result += 1

            return redirect(url_for("tasks", task_num=task.id+1))
    elif request.method == 'GET':
        if task_num == '0':
            return render_template('tasks.html', task = 0)
        else:
            get_tasks = session.query(Task).all()
            try:
                task = get_tasks[int(task_num)-1]
                listt = list(range(0, task.iterations+1))
                listt2 = list(range(0, task.sets))
                return render_template('tasks.html', task = task, listt=listt, listt2=listt2)
            except Exception:
                if name == '':
                    return render_template('tasks.html', task = 100, result = result, lenght = tasks_lenght, result_procent = (100/tasks_lenght)*result)
                newUser = User(name=name, result=result, result_procent=(100/tasks_lenght)*result, username=os.environ.get( "USERNAME" ))
                session.add(newUser)
                session.commit()
                return render_template('tasks.html', task = 100, result = result, lenght = tasks_lenght, result_procent = (100/tasks_lenght)*result)



def app_run():
    app.run(port=8081, host='127.0.0.1')

def game_run():
    os.system('C:\\Users\\User\\AppData\\Local\\Packages\\PythonSoftwareFoundation.Python.3.11_qbz5n2kfra8p0\\LocalCache\\local-packages\\Python311\\Scripts\\pygbag.exe C:\\Users\\User\\Desktop\\Petri_nets')



if __name__ == '__main__':
    process1 = multiprocessing.Process(target = app_run)
    process2 = multiprocessing.Process(target = game_run)
    process1.start()
    process2.start()














'''
Сети Петри - это математическая модель, 
используемая для описания и анализа параллельных 
и распределенных систем. Они применяются в различных 
областях, включая информационные технологии (моделирование 
и анализ производственных процессов, сети передачи данных), 
производственное управление (моделирование работы производственных линий), 
биологию (моделирование биохимических процессов в клетках) и другие области, 
где необходимо анализировать и управлять сложными системами.
'''